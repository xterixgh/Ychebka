import csv
import datetime
import glob
import os
from openpyxl import load_workbook


BASE = r"C:\Users\xteri\RiderProjects\YchebnayaPractika\_import_data"
OUT_DIR = r"C:\Users\xteri\RiderProjects\YchebnayaPractika\sql-export"
OUT_FILE = os.path.join(OUT_DIR, "seed_from_excel.sql")


def esc(value):
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def num(value):
    if value is None or str(value).strip() == "":
        return "NULL"
    try:
        x = float(value)
        if x.is_integer():
            return str(int(x))
        return str(x)
    except Exception:
        return esc(value)


def load_data():
    xlsx_files = glob.glob(BASE + r"\**\*.xlsx", recursive=True)
    csv_files = glob.glob(BASE + r"\**\*.csv", recursive=True)

    material_rows = []
    component_rows = []
    users = []

    for file_path in xlsx_files:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

        # Material file signature
        if "ГОСТ" in header and "Тип материала" in header:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                vals = list(row) + [None] * 9
                article = str(vals[0]).strip() if vals[0] is not None else ""
                name = str(vals[1]).strip() if vals[1] is not None else ""
                unit = str(vals[2]).strip() if vals[2] is not None else ""
                length = vals[3]
                qty = vals[4]
                material_type = str(vals[5]).strip() if vals[5] is not None else ""
                mass = vals[6]
                price = vals[7]
                gost = str(vals[8]).strip() if vals[8] is not None else None
                if not article or not name:
                    continue
                characteristic = f"Масса 1 м: {mass} кг" if mass is not None and str(mass).strip() else None
                material_rows.append((article, name, unit, qty, material_type, price, gost, length, characteristic))

        # Component file signature
        if "Вес" in header and "Единица измерения" in header and "Тип" in header:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                vals = list(row) + [None] * 7
                article = str(vals[0]).strip() if vals[0] is not None else ""
                name = str(vals[1]).strip() if vals[1] is not None else ""
                qty = vals[2]
                unit = str(vals[3]).strip() if vals[3] is not None else ""
                component_type = str(vals[4]).strip() if vals[4] is not None else ""
                price = vals[5]
                weight = vals[6]
                if not article or not name:
                    continue
                component_rows.append((article, name, unit, qty, component_type, price, weight))

    for file_path in csv_files:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as fp:
            reader = csv.reader(fp, delimiter=",")
            rows = list(reader)
        for row in rows[1:]:
            if len(row) < 5:
                continue
            surname = (row[0] or "").strip()
            name_pat = (row[1] or "").strip()
            login = (row[2] or "").strip()
            password = (row[3] or "").strip()
            role = (row[4] or "").strip()
            if not login:
                continue
            full_name = (surname + " " + name_pat).strip()
            users.append((login, password, full_name, role))

    return users, material_rows, component_rows


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    users, material_rows, component_rows = load_data()

    lines = []
    lines.append("-- Auto-generated from assignment spreadsheets")
    lines.append("-- Generated at: " + datetime.datetime.now().isoformat())
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    lines.append("-- Roles")
    lines.append('INSERT INTO "UserRole" ("RoleName") VALUES')
    lines.append("('Заказчик'),('Менеджер'),('Директор'),('Мастер'),('Конструктор')")
    lines.append('ON CONFLICT ("RoleName") DO NOTHING;')
    lines.append("")

    lines.append("-- Users from CSV")
    for login, password, full_name, role in users:
        lines.append('INSERT INTO "UserAccount" ("Login","PasswordHash","FullName","IdUserRole")')
        lines.append(f'SELECT {esc(login)}, {esc(password)}, {esc(full_name)}, ur."IdUserRole"')
        lines.append(f'FROM "UserRole" ur WHERE ur."RoleName" = {esc(role)}')
        lines.append('ON CONFLICT ("Login") DO NOTHING;')
    lines.append("")

    lines.append("-- Materials from Excel")
    for article, name, unit, qty, material_type, price, gost, length, characteristic in material_rows:
        lines.append(
            'INSERT INTO "Material" ("Article","MaterialName","UnitName","QuantityInStock","IdSupplier","Image","MaterialType","PurchasePrice","Gost","LengthValue","Characteristic") VALUES ('
            + ", ".join(
                [
                    esc(article),
                    esc(name),
                    esc(unit),
                    num(qty),
                    "NULL",
                    "NULL",
                    esc(material_type),
                    num(price),
                    esc(gost),
                    num(length),
                    esc(characteristic),
                ]
            )
            + ') ON CONFLICT ("Article") DO NOTHING;'
        )
    lines.append("")

    lines.append("-- Components from Excel")
    for article, name, unit, qty, component_type, price, weight in component_rows:
        lines.append(
            'INSERT INTO "Component" ("Article","ComponentName","UnitName","QuantityInStock","IdSupplier","Image","ComponentType","PurchasePrice","WeightValue") VALUES ('
            + ", ".join(
                [
                    esc(article),
                    esc(name),
                    esc(unit),
                    num(qty),
                    "NULL",
                    "NULL",
                    esc(component_type),
                    num(price),
                    num(weight),
                ]
            )
            + ') ON CONFLICT ("Article") DO NOTHING;'
        )
    lines.append("")
    lines.append("COMMIT;")

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print("Generated:", OUT_FILE)
    print("Users:", len(users))
    print("Materials:", len(material_rows))
    print("Components:", len(component_rows))


if __name__ == "__main__":
    main()
