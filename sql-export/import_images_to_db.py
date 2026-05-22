import glob
import os
from pathlib import Path

from openpyxl import load_workbook
import psycopg


DB_DSN = "host=localhost port=5432 dbname=BaseForYchebka user=postgres password=Bill0405h/"
BASE = Path(r"C:\Users\xteri\RiderProjects\YchebnayaPractika\_import_data\data\data")


def load_articles():
    materials_xlsx = BASE / "Материалы.xlsx"
    components_xlsx = BASE / "Комплектующие.xlsx"

    material_articles = set()
    component_articles = set()

    wb_m = load_workbook(materials_xlsx, data_only=True)
    ws_m = wb_m[wb_m.sheetnames[0]]
    for row in ws_m.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        material_articles.add(str(row[0]).strip())

    wb_c = load_workbook(components_xlsx, data_only=True)
    ws_c = wb_c[wb_c.sheetnames[0]]
    for row in ws_c.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        component_articles.add(str(row[0]).strip())

    return material_articles, component_articles


def get_all_image_files():
    exts = ("*.png", "*.jpg", "*.jpeg", "*.gif", "*.webp", "*.bmp")
    files = []
    for ext in exts:
        files.extend(glob.glob(str(BASE / "**" / ext), recursive=True))
    # include potential files without extension, like loginDEyua2018
    for f in glob.glob(str(BASE / "**" / "*"), recursive=True):
        p = Path(f)
        if p.is_file() and p.suffix == "":
            files.append(f)
    return [Path(f) for f in files]


def read_bytes(path: Path):
    with open(path, "rb") as fp:
        return fp.read()


def main():
    material_articles, component_articles = load_articles()
    files = get_all_image_files()

    users_updated = 0
    materials_updated = 0
    components_updated = 0

    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            for file_path in files:
                key = file_path.stem.strip()
                data = read_bytes(file_path)

                # 1) User photo by login (loginDE....)
                if key.startswith("login"):
                    cur.execute(
                        'UPDATE "UserAccount" SET "Photo" = %s WHERE "Login" = %s',
                        (data, key),
                    )
                    if cur.rowcount:
                        users_updated += cur.rowcount
                    continue

                # 2) Material image by article
                if key in material_articles:
                    cur.execute(
                        'UPDATE "Material" SET "Image" = %s WHERE "Article" = %s',
                        (data, key),
                    )
                    if cur.rowcount:
                        materials_updated += cur.rowcount
                    continue

                # 3) Component image by article
                if key in component_articles:
                    cur.execute(
                        'UPDATE "Component" SET "Image" = %s WHERE "Article" = %s',
                        (data, key),
                    )
                    if cur.rowcount:
                        components_updated += cur.rowcount
                    continue

        conn.commit()

    print("Image import completed:")
    print("  UserAccount.Photo updated:", users_updated)
    print("  Material.Image updated:", materials_updated)
    print("  Component.Image updated:", components_updated)


if __name__ == "__main__":
    main()
